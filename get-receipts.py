#  -*- coding: utf-8 -*-

##################################################
### Script by tbdk at github (Tobias Bonnesen) ###
##################################################

import requests
import json
import openpyxl
import datetime

def convert_date(date_string):
  # tager formattet: dd. monthinDanish yyyy 
  # smider følgende format ud: mm-dd-yy
  
  month_db = {
  "januar": "1", "februar": "2", "marts": "3", "april": "4", "maj": "5", "juni": "6", "juli": "7", 
  "august": "8", "september": "9", "oktober": "10", "november": "11", "december": "12"
  }

  split_string = date_string.split()
  day, month, year = split_string
  day = day[:-1] # removes the dot
  year = year
  month = month_db[month]

  date = "%s/%s/%s" % (month, day, year)
  return date

def get_user_id():
	"""Får fat i userID'et hvilket skal bruges i fremtidige requests"""

	url = "https://api.kvittering.dk/api/mobile/android/v3/login"
	headers = {
		'Host': 'api.kvittering.dk', 
		'Connection': 'Keep-Alive', 
		'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'
		}
	payload = "username=EMAILGOESHERE&password=RSAENCRYPTEDPASSWORDHERE" # Det er ikke selvforklarende hvordan man får fat i det. Undersøg de HTTPS kald der laves til
	# storebox for at se dit eget encrypterede password når du logger ind over mobil! (den skal være proxy'et så du kan aflytte fra din computer)
	# i overstående payload bruges min egen email og mit password i RSA enkrypteret form som jeg har fået fat i gennem aflytnning af mine egne requests til storebox. 
	# Passwordet er selvfølgelig fortroligt, udover at i har adgang til det gennem bilaget her.
	r = requests.post(url, headers=headers, data=payload)
	response = json.loads(r.text)
	userId = response["userId"]

	return userId

def get_receipt_ids():
	"Henter ID på alle bonnerne, hvilket bruges i requests når man vil have fat i én af dem."

	url = "https://api.kvittering.dk/api/mobile/android/v1/get-receipts?term=&offset=0&limit=25&orderBy=purchaseDateLocal_desc&directoryId=d7conxzqzu8hksfggphc6h4ed3vqvcqp&locale=en_GB&userId=%s" % (USER_ID)
	r = requests.get(url)
	receipt_list = json.loads(r.text)
	receipt_ids = []

	for receipt in receipt_list["receipts"]:
		receipt_ids.append(receipt["receiptId"])

	return receipt_ids

def create_receipt_list(receipt_ids):
	"Laver en liste med ID'erne på alle bonnerne"

	receipts = []
	
	for receipt_id in receipt_ids:
		
		url = "https://api.kvittering.dk/api/mobile/android/v1/get-receipt?receiptId=%s&userId=%s" % (receipt_id, USER_ID)
		r = requests.get(url)
		receipt = json.loads(r.text)
		receipts.append(receipt)
	return receipts

def create_excel():
	"Laver et excel dokument over alle ens bonner, med ugenr ud for hver vare, pris og antal."
	# tager formattet: mm-dd-yyyy

	wb = openpyxl.Workbook() # opens a new workbook.
	ws = wb.get_sheet_by_name('Sheet')
	ws.title = "Expenditure"
	ws2 = wb.create_sheet()
	ws2.title = "Item Database" 

	### Prepare sheet for population: ###

	ws['A1'] = "Store"
	ws['B1'] = "Week"
	ws['C1'] = "Item"
	ws['D1'] = "Price"

	row = 2
	total_price = 0
	num_of_weeks = []

	for receipt in RECEIPTS:
		date = convert_date(receipt["purchaseDate"])
		week = date
		month, day, year = week.split("/")
		
		week_number = datetime.date(int(year), int(month), int(day)).isocalendar()[1]
		store_name = receipt["merchantName"]
		total_price = total_price + float(receipt["totalOrderPriceValue"])

		num_of_weeks.append(week_number)

		for grocery in receipt["receiptLines"]:

			value_list = store_name, week_number, grocery["name"], float(grocery["itemPriceValue"])
			repeat = grocery["count"] 
			n = 0

			if repeat < 1: # fix for groceries that are weighted and have count < 1.
				repeat = 1

			while n != repeat:
				print "n is:", n, "repeat is:", repeat
				for index, value in enumerate(value_list):
					ws.cell(row=row, column=index+1).value = value
					print ws.cell(row=row, column=index+1)
				n = n + 1
				row = row + 1

	num_of_weeks_new = [week for week in num_of_weeks if week not in num_of_weeks]

	ws.cell(row= 1, column=8).value = "Summary:"
	ws.cell(row=2, column=8).value = "Num of weeks"
	ws.cell(row=2, column=9).value = len(num_of_weeks_new)
	ws.cell(row=3, column=8).value = "Total"
	ws.cell(row=3, column=9).value = "=SUM(D2:%s)" % ("D"+str(row - 1))
	ws.cell(row=4, column=8).value = "Weekly average"

	ws.cell(row=4, column=9).value = "=AVERAGE(I3:I2)"
	wb.save('ReceiptsNew.xlsx')

	
# Nedenstående er alle funktionskaldende så det sker. 
USER_ID = get_user_id()
RECEIPT_IDS = get_receipt_ids()
RECEIPTS = create_receipt_list(RECEIPT_IDS)
create_excel()

## Til trods for at det er min data i får adgang til, skal i være velkomne til at prøve at køre scriptet! 
